#!usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

list_grade = []
list_total = []

row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		list_grade.append([row_id, sum_v])
	row_id += 1
	
list_total = sorted(list_grade, key = lambda x:x[1], reverse=True)
a = int(len(list_total) * 0.3)
ap = int(a * 0.5)
b = int(len(list_total) * 0.7) - a
bp = int(b * 0.5)
c = int(len(list_total))- a -b
cp = int(c * 0.5)



for i in range(a):
	ws.cell(row = list_total[i][0], column = 8).value = 'A0'
for i in range(ap):
	ws.cell(row = list_total[i][0], column = 8).value = 'A+'
for i in range(a, a+b):
	ws.cell(row = list_total[i][0], column = 8).value = 'B0'
for i in range(a, a+bp):
	ws.cell(row = list_total[i][0], column = 8).value = 'B+'
for i in range(a+b, a+b+c):
	ws.cell(row = list_total[i][0], column = 8).value = 'C0'
for i in range(a+b, a+b+cp):
	ws.cell(row = list_total[i][0], column = 8).value = 'C+'

	
wb.save("student.xlsx")
wb.close()
		
		
			
